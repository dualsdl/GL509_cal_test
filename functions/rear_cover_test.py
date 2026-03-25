#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import numpy as np
import time
from datetime import datetime
from functions import util_yy

try:
    from stage_lib import ETEL
    import pysoslab_etel_stage as py_stage_etel
except ImportError:
    pass
import pysoslab_core
import pysoslab_user
import pysoslab_developer
import pysoslab_area
from stage_lib import DPIN

import cv2

# from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib as mpl
mpl.use("TkAgg")  # 예: TkAgg 백엔드 사용
mpl.rcParams['path.simplify'] = True
mpl.rcParams['path.simplify_threshold'] = 1.0


# #############  parameter start ###############
default_params = {
    "SERIAL_NAME": "COM4",
    "SERIAL_BAUDRATE": 921600,  # 921600, 115200
    "UDP_SENSOR_IP": "10.110.1.2",
    "UDP_SENSOR_PORT": 2000,
    "UDP_PC_IP": "10.110.1.3",
    "UDP_PC_PORT": 3000,
    "etel_stage_IP_addr": "10.110.1.200",
    "speed_mmps": 200.0,
    "target_angle": 90,  # OHT는 DG4090
    # "target_angle": -90,  # OBS는 92%
    "device_rotation_speed": 20.0,
    "LINEAR_STAGE_OFFSET": -43,
    # "LINEAR_STAGE_OFFSET_retro40": -43 + 188
    "test_area_path": "./OHT_filtering_validataion_area.json",
    # "test_area_path": "./empty_area.json",
    "empty_area_path": "./empty_area.json",

    "test_cond": {
        'test_path': [{
                'dist': 1300.0,
                'target_angle': -90,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90-35,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90+35,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': -90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90+35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90-35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': -45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90-35,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90+35,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': 45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90+35,
                'device_angle': 45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': 45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': 45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90-35,
                'device_angle': 45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': 45.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': 90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90-35,
                'device_angle': 90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90-35,
                'device_angle': 90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90+35,
                'device_angle': 90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 267.0,
                'target_angle': -90+35,
                'device_angle': 90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90,
                'device_angle': 90.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 1300.0,
                'target_angle': -90,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': True,
            },
            {
                'dist': 5000.0,
                'target_angle': -90,
                'device_angle': 0.0,
                'set_area': 30,
                'on_test': False,
            },
        ],
    },
}

# #############  parameter end ###############


def unittest() -> None:
    params = default_params
    # ETEL stage connect
    stage_etel = py_stage_etel.stage_etel()
    status = stage_etel.connect(params['etel_stage_IP_addr'], 3)
    if status:
        pass
    else:
        raise Exception('ETEL stage initialing FAIL')

    time.sleep(0.5)
    ETEL.search_home(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE)
    print('ROTARY_DEVICE home search done')

    ETEL.search_home(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET)
    print('LINEAR_TARGET home search done')
    stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, 5000.0, 400.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):
        time.sleep(0.2)

    ETEL.search_home(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET)
    print('ROTARY_TARGET home search done')

    # GL connect
    GL5_core = pysoslab_core.core()
    GL5_user = pysoslab_user.user()
    GL5_area = pysoslab_area.area()
    GL5_developer = pysoslab_developer.developer()
    # GL5_core.connectSerial(SERIAL_NAME, SERIAL_BAUDRATE)
    GL5_core.connectUDP(
        params['UDP_SENSOR_IP'],
        params['UDP_SENSOR_PORT'],
        params['UDP_PC_IP'],
        params['UDP_PC_PORT'],
        )

    input = dict()
    input['stage_etel'] = stage_etel
    input['GL5_core'] = GL5_core
    input['GL5_user'] = GL5_user
    input['GL5_developer'] = GL5_developer
    input['GL5_area'] = GL5_area

    stage_etel.setOffset(
        py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET,
        params['LINEAR_STAGE_OFFSET'],
    )
    
    dpin = DPIN.DpinStageHandler()
    dpin.connect("10.110.1.201", 184)
    dpin.searching_home()

    input['dpin'] = dpin

    data = run(input, params)

    test_name = 'rear_cover_test'
    path = f"./log/{test_name}"
    filename = f"{data['GL_serial']}_" + f'{data["test_time"]}' + f'{test_name}_rawdata'  # 확장자 없이
    util_yy.save_pickle_to_zip(data, path, filename)

    filename = f"{data['GL_serial']}_" + f'{data["test_time"]}' + f'{test_name}_report'  # 확장자 없이
    create_fail_report(data, f"{path}/{filename}.xlsx")


def data_acquisition(input: dict, params: dict = None) -> dict:
    output = dict()
    output['test_time'] = f"{datetime.now().year}" + \
                          f".{datetime.now().month}" + \
                          f".{datetime.now().day}" + \
                          f'_{datetime.now().strftime("%H.%M.%S")}'
    stage_etel = input['stage_etel']
    GL5_core = input['GL5_core']
    GL5_user = input['GL5_user']
    GL5_developer = input['GL5_developer']
    GL5_area = input['GL5_area']

    success, GL_serial = GL5_user.getSerialNum(GL5_core)
    if success:
        print(f"Serial_num = {GL_serial}")
    else:
        print("Unable to get a serial number")
    output['GL_serial'] = GL_serial

    success, areas = GL5_area.getAllAreaFromFile(params['test_area_path'])
    print(areas)
    if success:
        success = GL5_area.setAllAreaToSensor(GL5_core, GL5_user, areas)
        if success:
            print("Successfully set the areas")
        else:
            print("Unable to set the areas")
    else:
        print("Unable to read the areas from file")

    success = GL5_developer.setCompensation(GL5_core, 15)
    if success:
        print("Successfully set the compensation")
    else:
        print("Unable to set the compensation")

    success = GL5_developer.setLUTMode(GL5_core, 2)
    if success:
        print("Successfully set the lut_mode")
    else:
        print("Unable to set the lut_mode")

    send_str = 'area console setting'
    success, recv_str = GL5_developer.sendConsole(GL5_core, send_str)
    if success:
        print(f"send_str = {send_str}")
        print(f"recv_str = {recv_str}")
    else:
        print("Not responding to console messages")

    # LiDAR streaming start
    success = GL5_user.setStreamEnable(GL5_core, True)
    if success:
        print("Successfully enable data streaming")
    else:
        print("Failed to enable data streaming")

    init_dist = params['test_cond']['test_path'][0]['dist']
    init_angle = params['test_cond']['test_path'][0]['target_angle']
    init_device_angle = params['test_cond']['test_path'][0]['device_angle']
    
    # 초기 위치로 이동
    stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, init_dist, 400.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):
        time.sleep(0.2)

    stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET, init_angle, 30.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET):
        time.sleep(0.2)

    stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE, init_device_angle, params['device_rotation_speed'])
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE):
        time.sleep(0.2)

    output['fail_datas'] = []
    for meas_cond in params['test_cond']['test_path']:
        meas_dist = float(meas_cond['dist'])
        meas_angle = float(meas_cond['target_angle'])
        meas_device_angle = float(meas_cond['device_angle'])
        meas_set_area = int(meas_cond['set_area'])
        meas_on_test = bool(meas_cond['on_test'])

        etel_linear_spd = float(params['speed_mmps'])
        etel_device_spd = float(params['device_rotation_speed'])

        send_str = f'area set {meas_set_area}'
        success, recv_str = GL5_developer.sendConsole(GL5_core, send_str)        
        if success:
            print(f"send_str = {send_str}")
            print(f"recv_str = {recv_str}")
        else:
            print("Not responding to console messages")

        stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, meas_dist, etel_linear_spd)
        stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE, meas_device_angle, etel_device_spd)
        stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET, meas_angle, 30.0)
        
        logging_datas = []
        while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET) or \
            stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE) or \
            stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):

            success, frame_data = GL5_user.getLidarData(GL5_core, False)
            if not success:
                print("Failed to get a LiDAR data")
            else:
                logging_datas.append({
                    'frame_data': frame_data,
                    'test_dist': ETEL.get_current_position(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET),
                    'test_angle': ETEL.get_current_position(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET),
                    'test_device_angle': ETEL.get_current_position(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE),
                    })
        
        if meas_on_test:
            for data in logging_datas:
                fail_datas = dict()
                output_level = data['frame_data']['output_level']
                if output_level == 0:  # 미감지인 경우
                    fail_datas['test_dist'] = data['test_dist']
                    fail_datas['test_angle'] = data['test_angle']
                    fail_datas['test_device_angle'] = data['test_device_angle']
                    output['fail_datas'].append(fail_datas)
                    print(f"fail_datas = {fail_datas}")
                else:  # 감지인 경우
                    pass
    
    # 만약 output['fail_datas'] 가 비어있으면 테스트 통과
    if len(output['fail_datas']) == 0:
        output['result'] = 'PASS'
    else:
        output['result'] = 'FAIL'

    # output['fail_datas']를 output['fail_datas']['test_dist'] 기준으로 정렬
    output['fail_datas'] = sorted(output['fail_datas'], key=lambda x: x['test_dist'])

    stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, 5000.0, 400.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):
        time.sleep(0.2)

    stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE, 0.0, params['device_rotation_speed'])
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE):
        time.sleep(0.2)

    success, areas = GL5_area.getAllAreaFromFile(params['empty_area_path'])
    print(areas)
    if success:
        success = GL5_area.setAllAreaToSensor(GL5_core, GL5_user, areas)
        if success:
            print("Successfully set the areas")
        else:
            print("Unable to set the areas")
    else:
        print("Unable to read the areas from file")

    return output

def create_fail_report(result: dict, filename: str = 'fail_report.xlsx'):
    """
    1) 테스트 결과 요약 정보 (PASS/FAIL)를 '요약' 시트에 기록
    2) Fail 항목의 상세 내역을 '상세' 시트에 기록
    """
    import pandas as pd
    from openpyxl import load_workbook

    # 1) 요약 정보 작성
    summary_rows = [{
        '시리얼 넘버': result.get('GL_serial', 'UNKNOWN'),
        '테스트 이름': 'Rear Cover Test',
        '결과': result.get('result', 'UNKNOWN'),
        '테스트 시간': result.get('test_time', 'UNKNOWN'),             
    }]
    
    df_summary = pd.DataFrame(summary_rows)

    # 2) 상세 정보 작성 (Fail 항목만)
    detail_rows = []
    
    # 실패 데이터가 있는 경우에만 상세 정보 추가
    if 'fail_datas' in result and len(result['fail_datas']) > 0:
        for fail_data in result['fail_datas']:
            detail_rows.append({
                'Distance (mm)': fail_data.get('test_dist', None),
                'Target Angle (deg)': fail_data.get('test_angle', None),
                'Device Angle (deg)': fail_data.get('test_device_angle', None),
                'Status': 'Fail (미감지)'
            })
    
    df_details = pd.DataFrame(detail_rows)

    # 3) Excel 파일로 저장 (두 개의 시트)
    with pd.ExcelWriter(filename) as writer:
        df_summary.to_excel(writer, sheet_name='요약', index=False)
        df_details.to_excel(writer, sheet_name='상세', index=False)
    
    # 4) 자동 셀 크기 조정
    wb = load_workbook(filename)
    
    # 요약 시트의 셀 크기 조정
    summary_sheet = wb['요약']
    for col in range(1, len(df_summary.columns) + 2):  # +2는 인덱스와 마진을 위함
        col_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, len(df_summary) + 2):  # 헤더 포함
            cell_value = summary_sheet.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 4  # 여유 공간 추가
        summary_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # 상세 시트의 셀 크기 조정 및 소수점 형식 설정
    detail_sheet = wb['상세']
    for col in range(1, len(df_details.columns) + 2):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, len(df_details) + 2):  # 헤더 포함
            cell = detail_sheet.cell(row=row, column=col)
            cell_value = cell.value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
                
                # 숫자 서식 지정 (헤더 제외)
                if row > 1 and isinstance(cell_value, (int, float)):
                    from openpyxl.styles import numbers
                    cell.number_format = numbers.FORMAT_NUMBER_00  # 소수점 2자리 (예: 123.45)
                    
        adjusted_width = max_length + 4  # 여유 공간 추가
        detail_sheet.column_dimensions[col_letter].width = adjusted_width
    
    # 변경사항 저장
    wb.save(filename)

    print(f"리포트가 '{filename}' 파일로 저장되었습니다.")


def run(input: dict, params: dict = None) -> dict:
    stage_etel = input['stage_etel']

    stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, 5000.0, 400.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):
        time.sleep(0.2)

    stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET, 90.0, 30.0)  # 재귀반사판 설정
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET):
        time.sleep(0.2)

    from . import tx_level_test
    tx_level_test_parameter = {
        "scan_angles": ['0.0'],
        "OHT_target_angle": 90.0,
    }
    tx_level_rawdata, tx_level_output = tx_level_test.run(input, tx_level_test_parameter)
    tx_level = tx_level_output[f'{tx_level_test_parameter["scan_angles"][0]}']['tx_level_in_deg']
    # cv2.imshow('beam_height', tx_level_output['line_image'])
    # cv2.waitKey(1)
    input['dpin'].move_to_angle(tx_level)
    print('tx_level:', tx_level)

    if params is None:
        params = {}
    default_params.update(params)

    data = data_acquisition(input, default_params)

    return data


if __name__ == "__main__":
    unittest()
    # analysis_test()
