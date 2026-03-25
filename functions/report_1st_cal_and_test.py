from functions import util_yy
import numpy as np
from PySide6.QtWidgets import QApplication, QFileDialog
import sys
import cv2
from datetime import datetime


default_parameters = {
    'save_path': './A_ver_test_result/OBS/1st_cal_and_test/',
    'beam_size_log_path': './A_ver_test_result/OBS/beam_size_test/',
    'walk_error_log_path': './A_ver_test_result/OBS/walk_error/',
}


def create_excel_report(inputs, parameters):
        # Create Excel report
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        import os
        
        beam_size_result = inputs.get('beam_size_result')
        walkerror_result = inputs.get('walkerror_result')

        # 임시 이미지 파일 경로 저장을 위한 리스트
        temp_image_files = []

        # Create new workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "1st Cal & Test Report"
        
        # 기본 행 높이 설정
        DEFAULT_ROW_HEIGHT = 15
        
        # 픽셀을 포인트로 변환하는 상수 (1 픽셀 = 0.75 포인트)
        PIXEL_TO_POINT = 0.75
        
        # GL serial 검증
        if not inputs:
            raise ValueError("입력 데이터가 비어있습니다.")
            
        GL_serial = None
        for key in inputs.keys():
            if inputs[key] is not None:
                print(f"{key}의 GL 시리얼 번호: {inputs[key].get('GL_serial')}")

            if inputs[key] is None:
                continue
            elif GL_serial is None:
                GL_serial = inputs[key].get('GL_serial')
            elif inputs[key].get('GL_serial') is None:
                continue
            elif GL_serial != inputs[key].get('GL_serial'):
                for key in inputs.keys():
                    if inputs[key] is None:
                        continue
                    elif inputs[key].get('GL_serial') is not None:
                        print(f"{key}의 GL 시리얼 번호: {inputs[key].get('GL_serial')}")
                raise ValueError(f"{key}의 GL 시리얼 번호가 일치하지 않습니다: {GL_serial} != {inputs[key].get('GL_serial')}")
        
        if GL_serial is None:
            # 모든 GL_serial 출력
            raise ValueError("유효한 GL 시리얼 번호를 찾을 수 없습니다.")
        
        # Write headers
        ws['A1'] = 'GL Serial Number'
        ws['A2'] = GL_serial
        
        # Beam Size Results Section
        ws['A4'] = 'Beam Size Results'
        ws['A5'] = 'Parameter'
        ws['B5'] = 'Value'
        ws['A6'] = 'Test Time'
        ws['B6'] = beam_size_result.get('test_time', 'N/A') if beam_size_result else 'N/A'
        ws['A7'] = 'Major Axis (mm)'
        ws['B7'] = f"{beam_size_result.get('major_axis_length', 'N/A'):.3f}" if beam_size_result and beam_size_result.get('major_axis_length') is not None else 'N/A'
        ws['A8'] = 'Minor Axis (mm)'
        ws['B8'] = f"{beam_size_result.get('minor_axis_length', 'N/A'):.3f}" if beam_size_result and beam_size_result.get('minor_axis_length') is not None else 'N/A'
        ws['A9'] = 'Decenter Horizontal (deg)'
        ws['B9'] = f"{beam_size_result.get('decenter_horizontal_deg', 'N/A'):.3f}" if beam_size_result and beam_size_result.get('decenter_horizontal_deg') is not None else 'N/A'
        ws['A10'] = 'Decenter Vertical (deg)'
        ws['B10'] = f"{beam_size_result.get('decenter_vertical_deg', 'N/A'):.3f}" if beam_size_result and beam_size_result.get('decenter_vertical_deg') is not None else 'N/A'
        ws['A11'] = 'Area Fitting Percentage'
        ws['B11'] = f"{beam_size_result.get('area_fitting_percentage', 'N/A'):.3f}" if beam_size_result and beam_size_result.get('area_fitting_percentage') is not None else 'N/A'
        
        # Save beam size ellipse image and insert into Excel
        if beam_size_result and 'ellipse_img' in beam_size_result:
            img_path = os.path.join(parameters['save_path'], 'beam_size_ellipse.png')
            os.makedirs(os.path.dirname(img_path), exist_ok=True)
            cv2.imwrite(img_path, beam_size_result['ellipse_img'])
            temp_image_files.append(img_path)
            img = Image(img_path)
            ws.add_image(img, 'D4')
            # 이미지가 들어가고, 밑의 데이터가 들어가는 경우 이미지가 밑으로 벗어남
            # 이를 방지하기 위해 Beam Size Results Section 뒤에 빈 행 추가, 이미지 크기에 가변적으로 대응
            ws.insert_rows(12, 1)
            # Beam Size Results Section(A4~B11) 높이 계산
            beam_size_section_height = 0
            for row in range(4, 12):  # 4부터 11까지
                row_height = ws.row_dimensions[row].height if ws.row_dimensions[row].height is not None else DEFAULT_ROW_HEIGHT
                beam_size_section_height += row_height
            print(f"beam_size_section_height: {beam_size_section_height}")
            # 픽셀을 포인트로 변환하여 행 높이 설정
            ws.row_dimensions[12].height = img.height * PIXEL_TO_POINT - beam_size_section_height
        
        # Walk Error Results Section
        ws['A13'] = 'Walk Error Results'
        ws['A14'] = 'Parameter'
        ws['B14'] = 'Value'
        ws['A15'] = 'Test Time'
        ws['B15'] = walkerror_result.get('test_time', 'N/A') if walkerror_result else 'N/A'
        ws['A16'] = 'Distance Offset (mm)'
        ws['B16'] = f"{walkerror_result.get('distance_offset', 'N/A'):.3f}" if walkerror_result and walkerror_result.get('distance_offset') is not None else 'N/A'
        ws['A17'] = 'Std Residue Exceeds Threshold'
        ws['B17'] = 'Yes' if walkerror_result and walkerror_result.get('std_residue_exceeds_threshold') else 'N/A'
        ws['A18'] = 'Avg Residue Exceeds Threshold'
        ws['B18'] = 'Yes' if walkerror_result and walkerror_result.get('avg_residue_exceeds_threshold') else 'N/A'
        
        # Save walk error scatter plot and insert into Excel
        if walkerror_result and 'walk_error_graph' in walkerror_result:
            img_path = os.path.join(parameters['save_path'], 'walk_error_graph.png')
            os.makedirs(os.path.dirname(img_path), exist_ok=True)
            resized_img = cv2.resize(walkerror_result['walk_error_graph'], dsize=None, fx=0.75, fy=0.75, interpolation=cv2.INTER_AREA)
            cv2.imwrite(img_path, resized_img)
            temp_image_files.append(img_path)
            img = Image(img_path)

            ws.add_image(img, 'D13')
            # 이미지가 들어가고, 밑의 데이터가 들어가는 경우 이미지가 밑으로 벗어남
            # 이를 방지하기 위해 walk error result 뒤에 빈 행 추가, 이미지 크기에 가변적으로 대응
            ws.insert_rows(19, 1)
            # Walk Error Results Section(A13~B18) 높이 계산
            walk_error_section_height = 0
            for row in range(13, 19):  # 13부터 18까지
                row_height = ws.row_dimensions[row].height if ws.row_dimensions[row].height is not None else DEFAULT_ROW_HEIGHT
                walk_error_section_height += row_height
            # 픽셀을 포인트로 변환하여 행 높이 설정
            ws.row_dimensions[19].height = img.height * PIXEL_TO_POINT - walk_error_section_height

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2  # 여유 공간 추가
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the Excel file
        current_time = datetime.now().strftime("%Y%m%d_%H%M")
        report_path = os.path.join(parameters['save_path'], f"{current_time}_{GL_serial}_1st_cal_report.xlsx")
        wb.save(report_path)
        print(f"Report saved to: {report_path}")

        # 임시 이미지 파일 삭제
        for img_path in temp_image_files:
            try:
                if os.path.exists(img_path):
                    os.remove(img_path)
                    print(f"임시 이미지 파일 삭제됨: {img_path}")
            except Exception as e:
                print(f"이미지 파일 삭제 중 오류 발생: {str(e)}")


def run(inputs, parameters):
    try:
        beam_size_result = util_yy.load_pickle_from_zip(inputs['beam_size_result_fname'])
    except:
        beam_size_result = None
        
    try:
        walkerror_result = util_yy.load_pickle_from_zip(inputs['walkerror_result_fname'])
    except:
        walkerror_result = None

    report_inputs = {
        'beam_size_result': beam_size_result,
        'walkerror_result': walkerror_result,
    }
    
    try:
        # Call the report creation function
        create_excel_report(report_inputs, parameters)
    except ValueError as e:
        print(f"오류 발생: {str(e)}")
        return False
    return True


def unittest():
    app = QApplication(sys.argv)

    beam_size_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "파일 열기",                     # 대화상자 제목
        default_parameters['beam_size_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    walkerror_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "파일 열기",                     # 대화상자 제목
        default_parameters['walk_error_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )

    inputs = {
        'beam_size_result_fname': beam_size_result_fname,
        'walkerror_result_fname': walkerror_result_fname,
    }
    parameters = default_parameters

    run(inputs, parameters)
    app.quit()


if __name__ == '__main__':
    unittest()
    print('Done')