#!/usr/bin/env python3
#-*- coding : utf-8 -*-

import numpy as np
import cv2
from pypylon import pylon
import time
from datetime import datetime
import math
import os
from . import util_yy
import matplotlib.pyplot as plt

import pysoslab_etel_stage as py_stage_etel
import pysoslab_core
import pysoslab_user
import pysoslab_developer

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
# #############  parameter start ###############
default_parameters = {
    "SERIAL_NAME": "COM4",
    "SERIAL_BAUDRATE": 921600,  # 921600, 115200
    "UDP_SENSOR_IP": "10.110.1.2",
    "UDP_SENSOR_PORT": 2000,
    "UDP_PC_IP": "10.110.1.3",
    "UDP_PC_PORT": 3000,
    # "OHT_ld_high_voltage": 120,  # 15V
    # "OBS_ld_high_voltage": 150,  # 15V
    "OHT_target_angle": 90.0,
    "OBS_target_angle": -90.0,
    "DPIN_gonio_IP_addr": "10.110.1.201",
    "DPIN_gonio_port": 184,
    "etel_stage_IP_addr": "10.110.1.200",
    "LINEAR_STAGE_OFFSET": -43,
    "OHT_camera_exposure": 4000,
    "OBS_camera_exposure": 4000,
    "camera_FPS": 49.5,
    "camera_acqusition_num": 40,
    "image_ROI_x": [982, 1234],
    "image_ROI_y": [869, 1299],
    # "threshold_std": '1/e^2',  # '1/e^2' or 'FWHM'
    "threshold_std": 'FWHM',  # '1/e^2' or 'FWHM'
    "origin": [1109, 1118],  # 임시, (x, y) x는 좌우, y는 위아래
    # "px2mm_conversion_gain": 300 / (1106 - 840),  # d/px로 계산, 실제 측정한 거리d가 몇 픽셀px로 나오는가?
    "px2mm_conversion_gain": 1.145,  # d/px로 계산, 실제 측정한 거리d가 몇 픽셀px로 나오는가?
    "beamsize_pass_criteria": 25,  # mm
    "test_dist": 5000.0
}
# #############  parameter end ###############

# 전역 변수 (초기화)
mouse_x, mouse_y = -1, -1
bgr_text = ''


def generate_excel_report(report_data: dict, report_filename: str):
    """
    report_data 딕셔너리에는 다음 키들이 포함되어야 합니다:
      - 'GL_serial': GL 시리얼 번호 (문자열)
      - 'major_axis_length': 주축 길이 (mm, float)
      - 'minor_axis_length': 부축 길이 (mm, float)
      - 'decenter_horizontal_deg': 수평 decentered 각도 (deg, float)
      - 'decenter_vertical_deg': 수직 decentered 각도 (deg, float)
      - 'ellipse_img': ellipse 이미지 (numpy array, OpenCV 이미지)
      - (옵션) 'test_time': 테스트 시간 등 추가 정보

    report_filename: 생성할 Excel 파일 경로 (예: "report.xlsx")
    """
    # 1. 임시로 ellipse 이미지를 저장 (원본 해상도 유지)
    temp_img_filename = "temp_ellipse.png"
    ellipse_img = report_data.get('ellipse_img')
    if ellipse_img is None:
        raise ValueError("report_data에 'ellipse_img' 키가 필요합니다.")
    cv2.imwrite(temp_img_filename, ellipse_img)
    
    # 2. Excel Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "BeamSize Report"
    
    # 3. 기본 정보 입력 (원하는 위치나 서식을 자유롭게 변경 가능)
    ws["A1"] = "GL Serial"
    ws["B1"] = report_data.get("GL_serial", "N/A")
    
    ws["A2"] = "Test Time"
    ws["B2"] = report_data.get("test_time", "N/A")
    
    ws["A4"] = "Major Axis (mm)"
    ws["B4"] = report_data.get("major_axis_length", "N/A")
    
    ws["A5"] = "Minor Axis (mm)"
    ws["B5"] = report_data.get("minor_axis_length", "N/A")
    
    ws["A6"] = "Decentered Horizontal (deg)"
    ws["B6"] = report_data.get("decenter_horizontal_deg", "N/A")
    
    ws["A7"] = "Decentered Vertical (deg)"
    ws["B7"] = report_data.get("decenter_vertical_deg", "N/A")
    
    # 4. 이미지 삽입
    # 이미지 파일은 원본 해상도로 저장했으므로, openpyxl에서 별도의 스케일 조정 없이 그대로 삽입합니다.
    img = ExcelImage(temp_img_filename)
    # 이미지가 삽입될 셀 위치 지정 (예: "D1")
    ws.add_image(img, "D1")
    
    # 5. 엑셀 파일 저장
    wb.save(report_filename)
    
    # 6. 임시 이미지 파일 삭제 (필요 시)
    if os.path.exists(temp_img_filename):
        os.remove(temp_img_filename)

    print(f"Excel 리포트가 {report_filename}에 저장되었습니다.")


def data_acquisition(input: dict, params: dict = None) -> dict:
    outputs = dict()
    stage_etel = input['stage_etel']
    outputs['test_time'] = f"_{datetime.now().year}" + \
                          f".{datetime.now().month}" + \
                          f".{datetime.now().day}" + \
                          f'_{datetime.now().strftime("%H.%M.%S")}'
    stage_etel = input['stage_etel']
    GL5_core = input['GL5_core']
    GL5_user = input['GL5_user']
    GL5_developer = input['GL5_developer']

    success, GL_serial = GL5_user.getSerialNum(GL5_core)
    if success:
        print(f"Serial_num = {GL_serial}")
    else:
        print("Unable to get a serial number")
    outputs['GL_serial'] = GL_serial

    # 만약 Gl_serial의 4~5문자가 1W라면 OHT 센서, 1N이라면 OBS 센서
    model_name = GL_serial[4:6]
    print(f"Model name: {model_name}")
    if (model_name == '1W' or
            model_name == '1V'):
        target_angle = default_parameters['OHT_target_angle']
        camera_exposure = default_parameters['OHT_camera_exposure']
    elif (model_name == '1N' or
            model_name == '1M' or
            model_name == '1R' or
            model_name == '2N'):
        target_angle = default_parameters['OBS_target_angle']
        camera_exposure = default_parameters['OBS_camera_exposure']
    else:
        raise ValueError('Unknown sensor type')

    outputs['test_time'] = f"{datetime.now().year}" + \
                    f".{datetime.now().month}" + \
                    f".{datetime.now().day}" + \
                    f'_{datetime.now().strftime("%H.%M.%S")}'

    camera = pylon.InstantCamera(pylon.TlFactory.GetInstance().CreateFirstDevice())
    camera.Open()
    print("Using device ", camera.GetDeviceInfo().GetModelName())
    camera.ExposureTime = camera_exposure
    camera.AcquisitionFrameRateEnable.SetValue(True)
    camera.AcquisitionFrameRate = params['camera_FPS']
    fps = camera.ResultingFrameRate.GetValue()
    print(f'camera FPS is {fps:.3f}')

    # Grabing Continusely (video) with minimal delay
    camera.StartGrabbing(pylon.GrabStrategy_LatestImageOnly)
    converter = pylon.ImageFormatConverter()
    # converting to opencv bgr format
    converter.OutputPixelFormat = pylon.PixelType_BGR8packed
    converter.OutputBitAlignment = pylon.OutputBitAlignment_MsbAligned

    stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, 5000.0, 400.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):
        time.sleep(0.2)

    stage_etel.rotateTo(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET, target_angle, 30.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET):
        time.sleep(0.2)

    frame_cnt = 0
    imgs = []
    while camera.IsGrabbing():
        grabResult = camera.RetrieveResult(5000, pylon.TimeoutHandling_ThrowException)

        if grabResult.GrabSucceeded():
            # Access the image data
            image = converter.Convert(grabResult)
            img = image.GetArray()
            img = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
            imgs.append(img)

        frame_cnt = frame_cnt + 1
        if frame_cnt >= params['camera_acqusition_num']:
            # min_img = np.min(imgs, axis=0)  # back ground image without laser
            # max_img = np.max(imgs, axis=0)  # image with laser
            # var_img = max_img - min_img
            # outputs['variation_img'] = var_img
            # outputs['max_img'] = max_img
            # outputs['min_img'] = min_img
            outputs['raw_img'] = np.median(imgs, axis=0)
            break

    return outputs


def analysis(input:dict, params:dict = None) -> dict:
    output = dict()
    raw_img = input['raw_img']
    if len(raw_img.shape) == 3 and raw_img.shape[2] == 3:
        raw_img = cv2.cvtColor(raw_img, cv2.COLOR_BGR2GRAY)
    output['GL_serial'] = input['GL_serial']
    output['test_time'] = input['test_time']
    output['raw_img'] = raw_img.copy()

    # ROI crop
    cropped_img = raw_img[params['image_ROI_y'][0]:params['image_ROI_y'][1],
                          params['image_ROI_x'][0]:params['image_ROI_x'][1]]
    output['cropped_img'] = cropped_img.copy()

    # Convert to 8-bit if necessary
    if cropped_img.dtype != np.uint8:
        cropped_img = cv2.convertScaleAbs(cropped_img)

    # 스무딩 필터링
    k_size = 3
    filtered_img = cv2.bilateralFilter(cropped_img, k_size, 20, 20)
    output['filtered_img'] = filtered_img.copy()

    # 1/e^2 or FWHM 이진화
    back_ground_intensity = np.median(filtered_img.flatten())
    peak_intensity = np.max(filtered_img.flatten())
    diff_intensity = peak_intensity - back_ground_intensity
    if params['threshold_std'] == '1/e^2':
        binary_threshold = diff_intensity / np.e**2 + back_ground_intensity
    elif params['threshold_std'] == 'FWHM':
        binary_threshold = diff_intensity / 2 + back_ground_intensity
    else:
        raise ValueError('threshold_std should be 1/e^2 or FWHM')
    _, binary_img = cv2.threshold(filtered_img, binary_threshold, 255, cv2.THRESH_BINARY)
    output['binary_img'] = binary_img.copy()

    # 모폴로지 필터
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (25, 25))
    # moph_image = cv2.morphologyEx(binary_img, cv2.MORPH_OPEN, k)
    moph_img = cv2.morphologyEx(binary_img, cv2.MORPH_CLOSE, k)
    output['moph_img'] = moph_img.copy()

    # 레이블링
    # 객체 크기가 너무 크거나 작으면 에러처리 추가 필요
    _, labels, stats, _ = cv2.connectedComponentsWithStats(moph_img)
    s = stats[:, 4].argsort()[::-1]
    labels[np.where(labels != s[1])] = 0
    labels[np.where(labels == s[1])] = 255
    labels = np.uint8(labels)
    output['lebeling_img'] = labels.copy()

    # 타원 피팅
    contours, _ = cv2.findContours(labels, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if len(contours) > 0:
        largest_contour = max(contours, key=cv2.contourArea)
        if len(largest_contour) >= 5:  # fitEllipse requires at least 5 points
            ellipse = cv2.fitEllipse(largest_contour)
            output['ellipse'] = ellipse

            # Draw the ellipse on the original image
            cv2.ellipse(raw_img, ellipse, (255, 0, 0), 2)  # Red color with thickness 2

            # Extract and print the major and minor axis lengths
            (center, axes, orientation) = ellipse
            major_axis_length = max(axes)
            minor_axis_length = min(axes)
            # print(f"Major axis length: {major_axis_length:.2f} pixels")
            # print(f"Minor axis length: {minor_axis_length:.2f} pixels")        

            ellipse_img = cv2.cvtColor(cropped_img, cv2.COLOR_GRAY2BGR).copy()
            text = f"Major: {major_axis_length:.2f}px, Minor: {minor_axis_length:.2f}px"
            cv2.putText(ellipse_img, text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)
            cv2.ellipse(ellipse_img, ellipse, (0, 255, 0), 1)  # Draw ellipse in green color
            output['major_axis_length'] = major_axis_length * params['px2mm_conversion_gain']
            output['minor_axis_length'] = minor_axis_length * params['px2mm_conversion_gain']
            text = f"Major: {output['major_axis_length']:.2f}mm, Minor: {output['minor_axis_length']:.2f}mm"
            cv2.putText(ellipse_img, text, (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)
            # 이진화 영역 색 주황색을 더하여 표시
            # ellipse_image[np.where(labels == 255)] = ellipse_image[np.where(labels == 255)]+np.array([0, 0, 50])
            ellipse_img[np.where(labels == 255)] = np.clip(ellipse_img[np.where(labels == 255)]+np.array([0, 0, 80]), 0, 255)
            # 좌측 상단에 legend 추가
            cv2.putText(ellipse_img, "laser area", (10, 10), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)
            cv2.rectangle(ellipse_img, (62, 6), (122, 11), (50, 50, 150), -1)  # (왼쪽 위 좌표, 오른쪽 아래 좌표, 색, 두께)

            # origin 포인트 추가 (원점 좌표를 이미지 ROI에 맞게 변환)
            # 원래 origin은 (x, y) 좌표라고 가정하며, cropped_img는 
            # raw_img[image_ROI_x[0]:image_ROI_x[1], image_ROI_y[0]:image_ROI_y[1]]로 생성됨.
            # 따라서 로컬 좌표는 (origin_x - image_ROI_y[0], origin_y - image_ROI_x[0])
            origin_local = (params['origin'][0] - params['image_ROI_y'][0], params['origin'][1] - params['image_ROI_x'][0])
            # cv2.circle(ellipse_img, origin_local, 3, (0, 0, 255), -1)  # 빨간색 점
            # 2) 십자선 길이(px) 설정
            cross_size = 30
            # 4) 가로선: (x - cross_size, y) -> (x + cross_size, y)
            start_point = (origin_local[0] - cross_size, origin_local[1])
            end_point = (origin_local[0] + cross_size, origin_local[1])
            util_yy.draw_dashed_line(ellipse_img, start_point, end_point)

            # 5) 세로선: (x, y - cross_size) -> (x, y + cross_size)
            start_point = (origin_local[0], origin_local[1] - cross_size)
            end_point = (origin_local[0], origin_local[1] + cross_size)
            util_yy.draw_dashed_line(ellipse_img, start_point, end_point)

            dx_pixels = center[0] - origin_local[0]
            dy_pixels = center[1] - origin_local[1]
            dx_mm = dx_pixels * params['px2mm_conversion_gain']
            dy_mm = dy_pixels * params['px2mm_conversion_gain']
            dx_deg = math.degrees(math.atan(dx_mm / params['test_dist']))
            dy_deg = math.degrees(math.atan(dy_mm / params['test_dist']))

            output['center_dx_pixels'] = dx_pixels
            output['center_dy_pixels'] = dy_pixels
            output['decenter_vertical_deg'] = dy_deg
            output['decenter_horizontal_deg'] = dx_deg
               
            # 계산된 x, y 축 차이를 이미지에 텍스트로 추가 (표시 위치는 필요에 따라 조정)
            diff_text = f"d_hor: {dx_deg:.2f}deg, d_ver: {dy_deg:.2f}deg"
            cv2.putText(ellipse_img, diff_text, (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)

            # (1) 면적 기반 피팅 퍼센트 구하기
            # ------------------------------
            # 1) 컨투어 마스크
            contour_mask = labels  # 이미 0/255 형태라면 이대로 사용
            h, w = contour_mask.shape[:2]

            # 2) 엘립스 마스크 그리기
            ellipse_mask = np.zeros((h, w), dtype=np.uint8)

            cv2.ellipse(
                ellipse_mask, ellipse,
                255,  # 흰색
                -1    # 내부 채움
            )

            # 3) Intersection & Union
            intersection = cv2.bitwise_and(contour_mask, ellipse_mask)
            union = cv2.bitwise_or(contour_mask, ellipse_mask)

            inter_area = cv2.countNonZero(intersection)
            union_area = cv2.countNonZero(union)

            if union_area > 0:
                area_fitting_percentage = (inter_area / union_area) * 100.0
            else:
                area_fitting_percentage = 0.0

            text = f"area_fitting_percentage: {area_fitting_percentage:.2f}%"
            cv2.putText(ellipse_img, text, (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 255, 255), 1)
            output['area_fitting_percentage'] = area_fitting_percentage
            output['ellipse_img'] = ellipse_img

        else:
            print("Not enough points to fit an ellipse.")
    else:
        print("No contours found.")

    # 결과 판정, 기준 확인 및 업데이트 필요
    if 'major_axis_length' in output.keys():
        if output['major_axis_length'] < params['beamsize_pass_criteria']:
            output['result'] = 'PASS'
        else:
            output['result'] = 'FAIL'
    else:
        output['result'] = 'FAIL'

    # debug image show
    # 윈도우 생성 + 마우스 콜백 등록
    debug = False
    # debug = True
    if debug:
        cv2.namedWindow('Debug')
        debug_img = ellipse_img
        cv2.setMouseCallback('Debug', mouse_callback, debug_img)

        while True:
            # 상태바 영역이 붙은 형태로 이미지 표시
            show_debug_image_with_bar(debug_img)

            key = cv2.waitKey(1) & 0xFF
            if key == 27:  # ESC
                break

        cv2.destroyAllWindows()

    return output


def show_debug_image_with_bar(img):
    """
    img: 흑백(단일 채널) 또는 컬러 이미지라 가정.
    """
    global mouse_x, mouse_y, pixel_text

    # 1) 상태바 높이 지정
    status_bar_height = 30

    # 2) 아래쪽에 상태바 공간을 만든 새 캔버스 생성
    h, w = img.shape[:2]
    display_img = np.zeros((h + status_bar_height, w, 3), dtype=np.uint8)

    # 3) 상단 영역에 원본 이미지를 복사
    if len(img.shape) == 2:  # Grayscale image
        img_color = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    else:  # Color image
        img_color = img

    display_img[:h, :w] = img_color

    # 4) 하단 바 영역을 회색 사각형으로 채우기
    cv2.rectangle(display_img,
                  (0, h),            # 시작점 (왼쪽 위)
                  (w, h + status_bar_height),  # 끝점 (오른쪽 아래)
                  (50, 50, 50),      # 회색 (B,G,R)
                  -1)                # -1 -> 내부 전체 채움

    # 5) 좌표 및 픽셀값 텍스트 만들기
    if 0 <= mouse_x < w and 0 <= mouse_y < h:
        if len(img.shape) == 2:  # Grayscale image
            gray_val = img[mouse_y, mouse_x]
            pixel_text = f"Gray={gray_val}"
        else:  # Color image
            b, g, r = img[mouse_y, mouse_x]
            pixel_text = f"BGR=({b},{g},{r})"
        text = f"pos=({mouse_x},{mouse_y}), {pixel_text}"
    else:
        text = "Out of image bounds."

    # 6) 하단 바 영역에 텍스트 그리기
    cv2.putText(display_img,
                text,
                (10, h + 20),  # 바닥에서 10px 정도 띄워서 표시
                cv2.FONT_HERSHEY_SIMPLEX,
                0.3,           # 글자 크기
                (255, 255, 255),   # 흰색
                1, cv2.LINE_AA)

    # 7) 완성된 이미지를 imshow로 표시
    cv2.imshow('Debug', display_img)


# 마우스 이벤트 처리 함수
def mouse_callback(event, x, y, flags, param):
    global mouse_x, mouse_y, pixel_text
    # param: 디버그용(시각화)으로 쓸 원본 이미지(혹은 ROI)

    if event == cv2.EVENT_MOUSEMOVE:
        mouse_x, mouse_y = x, y
        # 이미지 범위 안이라면 픽셀값 추출
        if 0 <= y < param.shape[0] and 0 <= x < param.shape[1]:
            # 만약 원본이 흑백(1채널)이면 param[y,x] 자체가 숫자 하나
            if len(param.shape) == 2:
                # 그레이스케일
                gray_val = param[y, x]
                pixel_text = f"Gray={gray_val}"
            else:
                # 컬러(BGR)
                b, g, r = param[y, x]
                # pixel_text = f"Gray={b}"
                pixel_text = f"BGR=({b},{g},{r})"
        else:
            pixel_text = ''


def test_from_raw_data():
    # Load raw data from pickle file
    import util_yy
    raw_data = util_yy.load_pickle_from_zip('./log/beam_size_test/G507VxxYY-T1_2025.3.11_17.15.57beam_size_test_rawdata.zip')
    
    # Run analysis
    result = analysis(raw_data, default_parameters)
    
    # Save output images
    cv2.imwrite('./beam_size_test_data/log/1.cropped_img.png', result['cropped_img'])
    cv2.imwrite('./beam_size_test_data/log/2.filtered_img.png', result['filtered_img']) 
    cv2.imwrite('./beam_size_test_data/log/3.binary_img.png', result['binary_img'])
    cv2.imwrite('./beam_size_test_data/log/4.moph_img.png', result['moph_img'])
    cv2.imwrite('./beam_size_test_data/log/5.lebeling_img.png', result['lebeling_img'])
    cv2.imwrite('./beam_size_test_data/log/6.ellipse_img.png', result['ellipse_img'])
    
    test_name = 'beam_size_test'
    path = f"./log/{test_name}"

    filename = f"{raw_data['GL_serial']}_" + f'{raw_data["test_time"]}' + f'{test_name}_procdata'  # 확장자 없이
    util_yy.save_pickle_to_zip(result, path, filename)


    # save image    
    filename = f"{raw_data['GL_serial']}_" + f'{raw_data["test_time"]}' + f'{test_name}'  # 확장자 없이
    cv2.imwrite(f"{path}/{filename}_beam_size.png", result['ellipse_img'])

    report_data = {
        "GL_serial": result.get("GL_serial", "UNKNOWN"),
        "test_time": result.get("test_time", "N/A"),
        "major_axis_length": result.get("major_axis_length", 0),
        "minor_axis_length": result.get("minor_axis_length", 0),
        "decenter_horizontal_deg": result.get("decenter_horizontal_deg", 0),
        "decenter_vertical_deg": result.get("decenter_vertical_deg", 0),
        "ellipse_img": result.get("ellipse_img")  # OpenCV 이미지
    }
    generate_excel_report(report_data, f"{path}/{filename}_beam_size.xlsx")


def run(inputs:dict, params:dict = None) -> dict:
    default_parameters.update(params)

    GL5_core = inputs['GL5_core']
    GL5_user = inputs['GL5_user']
    GL5_developer = inputs['GL5_developer']

    success, GL_serial = GL5_user.getSerialNum(GL5_core)
    if success:
        print(f"Serial_num = {GL_serial}")
    else:
        print("Unable to get a serial number")

    # GL 원포인트 셋
    success = GL5_developer.setOperationMode(GL5_core, 1)
    if success:
        print("Successfully set the Setup Mode")
    else:
        print("Unable to set the Setup Mode")

    # data acquisition
    raw_data = data_acquisition(inputs, default_parameters)
    # analysis
    output = analysis(raw_data, default_parameters)

    return raw_data, output


def unittest():

    # LiDAR Serial number read    
    GL5_core = pysoslab_core.core()
    GL5_user = pysoslab_user.user()
    GL5_developer = pysoslab_developer.developer()
    # GL5_core.connectSerial(SERIAL_NAME, SERIAL_BAUDRATE)
    GL5_core.connectUDP(default_parameters['UDP_SENSOR_IP'],
                        default_parameters['UDP_SENSOR_PORT'],
                        default_parameters['UDP_PC_IP'],
                        default_parameters['UDP_PC_PORT'])
    success, GL_serial = GL5_user.getSerialNum(GL5_core)
    if success:
        print(f"Serial_num = {GL_serial}")
    else:
        print("Unable to get a serial number")

    # DPIN stage connect
    # stage_dpin = DpinStageHandler()
    # stage_dpin.connect(DPIN_gonio_IP_addr, DPIN_gonio_port)
    # stage_dpin.searching_home()
    # stage_dpin.move_to_angle(0.0)

    # ETEL stage connect
    stage_etel = py_stage_etel.stage_etel()
    status = stage_etel.connect(default_parameters['etel_stage_IP_addr'], 3)
    if status:
        pass
    else:
        raise Exception('ETEL stage initialing FAIL')

    stage_etel.setOffset(
        py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, default_parameters['LINEAR_STAGE_OFFSET']
    )

    time.sleep(0.5)
    ETEL.search_home(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.ROTARY_DEVICE)
    print('ROTARY_DEVICE home search done')
    ETEL.search_home(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET)
    print('LINEAR_TARGET home search done')
    stage_etel.moveTo(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET, default_parameters['test_dist'], 400.0)
    while stage_etel.isMoving(py_stage_etel.STAGE_AXIS_INDEX.LINEAR_TARGET):
        time.sleep(0.2)
    ETEL.search_home(stage_etel, py_stage_etel.STAGE_AXIS_INDEX.ROTARY_TARGET)
    print('ROTARY_TARGET home search done')

    input = dict()
    input['stage_etel'] = stage_etel
    input['GL5_core'] = GL5_core
    input['GL5_user'] = GL5_user
    input['GL5_developer'] = GL5_developer

    raw_data, result = run(input, default_parameters)

    test_name = 'beam_size_test'
    path = f"./log/{test_name}"
    filename = f"{raw_data['GL_serial']}_" + f'{raw_data["test_time"]}' + f'{test_name}_rawdata'  # 확장자 없이
    util_yy.save_pickle_to_zip(raw_data, path, filename)

    filename = f"{raw_data['GL_serial']}_" + f'{raw_data["test_time"]}' + f'{test_name}_procdata'  # 확장자 없이
    util_yy.save_pickle_to_zip(result, path, filename)


    # save image    
    filename = f"{raw_data['GL_serial']}_" + f'{raw_data["test_time"]}' + f'{test_name}'  # 확장자 없이
    cv2.imwrite(f"{path}/{filename}_beam_size.png", result['ellipse_img'])

    report_data = {
        "GL_serial": result.get("GL_serial", "UNKNOWN"),
        "test_time": result.get("test_time", "N/A"),
        "major_axis_length": result.get("major_axis_length", 0),
        "minor_axis_length": result.get("minor_axis_length", 0),
        "decenter_horizontal_deg": result.get("decenter_horizontal_deg", 0),
        "decenter_vertical_deg": result.get("decenter_vertical_deg", 0),
        "ellipse_img": result.get("ellipse_img")  # OpenCV 이미지
    }
    generate_excel_report(report_data, f"{path}/{filename}_beam_size.xlsx")


if __name__ == "__main__":    
    from ETEL_stage_python310.gwangju_motorized_stage import test_stage as ETEL
    # test_from_raw_data()
    unittest()
