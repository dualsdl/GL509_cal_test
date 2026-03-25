
try:
    from functions import util_yy
except:
    import util_yy
import numpy as np
import pandas as pd
from PySide6.QtWidgets import QApplication, QFileDialog
import sys
import cv2
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


default_parameters = {
    'save_path': './A_ver_test_result/2nd_cal_and_test/',
    'tx_level_log_path': './A_ver_test_result/tx_level/',
    'home_position_log_path': './A_ver_test_result/home_position/',
    'distance_test_log_path': './A_ver_test_result/distance_test/',
    'oht_filtering_table_log_path': './A_ver_test_result/oht_filtering_table/',
    'oht_filtering_validation_log_path': './A_ver_test_result/oht_filtering_validation/',
    'max_dist_log_path': './A_ver_test_result/max_dist/',
    'beam_size_log_path': './A_ver_test_result/beam_size_test/',
    'walk_error_log_path': './A_ver_test_result/log/walk_error/',
    'rear_cover_test_path': './log/rear_cover_test/',
}


def create_excel_report(inputs, parameters):
        # Create Excel report
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        import os
        
        tx_level_result = inputs.get('tx_level_result')
        home_position_result = inputs.get('home_position_result')
        distance_test_result = inputs.get('distance_test_result')
        oht_filtering_table_result = inputs.get('oht_filtering_table_result')
        oht_filtering_validation_result = inputs.get('oht_filtering_validation_result')
        max_dist_result = inputs.get('max_dist_result')

        # 임시 이미지 파일 경로 저장을 위한 리스트
        temp_image_files = []

        # Create new workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "2nd Cal & Test Report"
        
        # 기본 행 높이 설정
        DEFAULT_ROW_HEIGHT = 15
        
        # 픽셀을 포인트로 변환하는 상수 (1 픽셀 = 0.75 포인트)
        PIXEL_TO_POINT = 0.75
        
        # GL serial 검증
        if not inputs:
            raise ValueError("입력 데이터가 비어있습니다.")
            
        GL_serial = None
        for key in inputs.keys():
            # if inputs[key] is not None:
            #     print(f"{key}의 GL 시리얼 번호: {inputs[key].get('GL_serial')}")

            if inputs[key] is None:
                continue
            elif GL_serial is None:
                GL_serial = inputs[key].get('GL_serial')
            elif inputs[key].get('GL_serial') is None:
                continue
            elif GL_serial != inputs[key].get('GL_serial'):
                for key in inputs.keys():
                    if inputs[key] is None:
                        continue
                    elif inputs[key].get('GL_serial') is not None:
                        print(f"{key}의 GL 시리얼 번호: {inputs[key].get('GL_serial')}")
                raise ValueError(f"{key}의 GL 시리얼 번호가 일치하지 않습니다: {GL_serial} != {inputs[key].get('GL_serial')}")
        
        if GL_serial is None:
            # 모든 GL_serial 출력
            raise ValueError("유효한 GL 시리얼 번호를 찾을 수 없습니다.")
        
        # Write headers
        ws['A1'] = 'GL Serial Number'
        ws['A2'] = GL_serial

        # TX Level Results Section (시작: A4)
        ws['A4'] = 'TX Level Results'
        ws['A5'] = 'Parameter'
        ws['B5'] = 'Value'
        ws['A6'] = 'Test Time'
        ws['B6'] = tx_level_result.get('test_time', 'N/A') if tx_level_result else 'N/A'
        ws['A7'] = 'TX Level Test Result'
        ws['B7'] = 'PASS' if tx_level_result and tx_level_result.get('tx_level_Pass/Fail', False) else 'FAIL'

        #-135deg tx level, 
        ws['A8'] = '-135deg tx level[deg]'
        ws['B8'] = f"{tx_level_result.get('-135.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('-135.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        #-90deg tx level
        ws['A9'] = '-90deg tx level[deg]'
        ws['B9'] = f"{tx_level_result.get('-90.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('-90.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        #-45deg tx level
        ws['A10'] = '-45deg tx level[deg]'
        ws['B10'] = f"{tx_level_result.get('-45.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('-45.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        #0deg tx level
        ws['A11'] = '0deg tx level[deg]'
        ws['B11'] = f"{tx_level_result.get('0.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('0.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        #45deg tx level
        ws['A12'] = '45deg tx level[deg]'
        ws['B12'] = f"{tx_level_result.get('45.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('45.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        #90deg tx level
        ws['A13'] = '90deg tx level[deg]'
        ws['B13'] = f"{tx_level_result.get('90.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('90.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        #135deg tx level
        ws['A14'] = '135deg tx level[deg]'
        ws['B14'] = f"{tx_level_result.get('135.0', {}).get('tx_level_in_deg', 'N/A'):.2f}" if tx_level_result and tx_level_result.get('135.0', {}).get('tx_level_in_deg') is not None else 'N/A'
        
        if tx_level_result and 'line_image' in tx_level_result:
            img_path = os.path.join(parameters['save_path'], 'tx_level_line.png')
            os.makedirs(os.path.dirname(img_path), exist_ok=True)
            resized_img = cv2.resize(tx_level_result['line_image'], dsize=None, fx=0.75, fy=0.75, interpolation=cv2.INTER_AREA)

            cv2.imwrite(img_path, resized_img)
            temp_image_files.append(img_path)
            img = Image(img_path)
            ws.add_image(img, 'D4')
            
            ws.insert_rows(15, 1)
            tx_level_section_height = 0
            for row in range(4, 15):
                row_height = ws.row_dimensions[row].height if ws.row_dimensions[row].height is not None else DEFAULT_ROW_HEIGHT
                tx_level_section_height += row_height
            ws.row_dimensions[15].height = img.height * PIXEL_TO_POINT - tx_level_section_height
        
        # Home Position Results Section (시작: A17)
        ws['A17'] = 'Home Position Results'
        ws['A18'] = 'Parameter'
        ws['B18'] = 'Value'
        ws['A19'] = 'Test Time'
        ws['B19'] = home_position_result.get('test_time', 'N/A') if home_position_result else 'N/A'
        ws['A20'] = 'Home Position Test Result'
        ws['B20'] = home_position_result.get('home_position_test_result', 'FAIL') if home_position_result else 'N/A'
        ws['A21'] = 'Home Position Setting'
        ws['B21'] = f"{home_position_result.get('prev_home_position', 'N/A')}"  if home_position_result else 'N/A'
        ws['A22'] = 'Home Position Error (idx)'
        ws['B22'] = f"{home_position_result.get('home_position_error', 'N/A'):.3f}" if home_position_result else 'N/A'
        
        # Distance Offset Results Section (시작: A24)
        ws['A24'] = 'Distance Test Results'
        ws['A25'] = 'Parameter'
        ws['B25'] = 'Value'
        ws['A26'] = 'Test Time'
        ws['B26'] = distance_test_result.get('test_time', 'N/A') if distance_test_result else 'N/A'
        ws['A27'] = 'back_reflector_distance_target (mm)'
        ws['B27'] = f"{distance_test_result.get('back_reflector_distance_target', 'N/A'):.3f}" if distance_test_result and distance_test_result.get('back_reflector_distance_target') is not None else 'N/A'
        ws['A28'] = 'Accuracy'
        if distance_test_result and 'results' in distance_test_result:
            for i in range(len(distance_test_result['results'])):
                if distance_test_result['results'][i]['accuracy_pass'] == 'PASS':
                    accuracy_pass = 'PASS'
                else:
                    accuracy_pass = 'FAIL'
                    break
            ws['B28'] = accuracy_pass
        else:
            ws['B28'] = 'N/A'
            
        ws['A29'] = 'Precision'
        if distance_test_result and 'results' in distance_test_result:
            for i in range(len(distance_test_result['results'])):
                if distance_test_result['results'][i]['precision_pass'] == 'PASS':
                    precision_pass = 'PASS'
                else:
                    precision_pass = 'FAIL'
                    break
            ws['B29'] = precision_pass
        else:
            ws['B29'] = 'N/A'

        # 두번째 sheet를 distance_test_result로 만들고
        ws2 = wb.create_sheet('Dist test')        
        # DataFrame 생성 및 데이터 쓰기
        if distance_test_result and 'results' in distance_test_result:
            df = pd.DataFrame(distance_test_result['results'])        
            # precision, accuracy 열만 소수점 4자리로 포맷팅
            for col in ['precision', 'accuracy']:
                df[col] = df[col].apply(lambda x: f"{x:.3f}")            
            for r in dataframe_to_rows(df, index=False, header=True):
                ws2.append(r)            
            # Setting column widths
            for column in ws2.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass
                adjusted_width = (max_length + 2)
                ws2.column_dimensions[column_name].width = adjusted_width
            # Setting header styles  
            header_font = Font(bold=True)
            for cell in ws2["1:1"]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')            
            # FAIL이 있는 행에 배경색 옅은 분홍색 넣기
            light_pink_fill = PatternFill(start_color='FFD1D1', end_color='FFD1D1', fill_type='solid')
            for row in ws2.iter_rows(min_row=2):  # 헤더 제외하고 시작
                for cell in row:
                    if 'FAIL' in str(cell.value):
                        for c in row:
                            c.fill = light_pink_fill
                        break
        else:
            ws2['A1'] = 'No Data Available'
        
        # OHT Filtering Table Results Section (시작: A31)
        ws['A31'] = 'OHT Filtering Table Results'        
        ws['A32'] = 'Parameter'
        ws['B32'] = 'Value'
        ws['A33'] = 'Test Time'
        ws['B33'] = oht_filtering_table_result.get('test_time', 'N/A') if oht_filtering_table_result else 'N/A'

        GL_setting = oht_filtering_table_result.get('GL_setting', {}) if oht_filtering_table_result else {}
        ws['A34'] = 'back_reflector_pulse_width_target'
        ws['B34'] = f"{GL_setting.get('back_reflector_pulse_width_target', 'N/A')}"
        ws['A35'] = 'PD High Voltage'
        ws['B35'] = f"{GL_setting.get('pd_high_voltage', 'N/A')}"

        # 3번째 시트를 추가하고, 이름을 'OHT Filtering Table'로 변경
        ws3 = wb.create_sheet('OHT LUT')
        if oht_filtering_table_result and 'oht_table' in oht_filtering_table_result:
            # NumPy 배열을 DataFrame으로 변환
            df_oht = pd.DataFrame(oht_filtering_table_result['oht_table'])
            # 모든 숫자를 정수로 변환
            df_oht = df_oht.astype(int)
            for row in dataframe_to_rows(df_oht, index=False, header=False):
                ws3.append(row)
        else:
            ws3['A1'] = 'No Data Available'

        # 5m 이미지 추가
        if oht_filtering_table_result and 'img_5m' in oht_filtering_table_result:
            img_path = os.path.join(parameters['save_path'], 'oht_filtering_table.png')
            os.makedirs(os.path.dirname(img_path), exist_ok=True)
            # BGR에서 RGB로 변환
            img_rgb = cv2.cvtColor(oht_filtering_table_result['img_5m'], cv2.COLOR_BGR2RGB)
            img_rgb = cv2.resize(img_rgb, dsize=None, fx=0.65, fy=0.65, interpolation=cv2.INTER_AREA)

            cv2.imwrite(img_path, img_rgb)
            temp_image_files.append(img_path)
            img = Image(img_path)
            ws.add_image(img, 'D31')
            
            ws.insert_rows(36, 1)
            tx_level_section_height = 0
            for row in range(30, 36):
                row_height = ws.row_dimensions[row].height if ws.row_dimensions[row].height is not None else DEFAULT_ROW_HEIGHT
                tx_level_section_height += row_height
            ws.row_dimensions[36].height = img.height * PIXEL_TO_POINT - tx_level_section_height

        # OHT Filtering Validation Results Section (시작: A37)
        ws['A37'] = 'OHT Filtering Validation Results'
        ws['A38'] = 'Parameter'
        ws['B38'] = 'Value'
        ws['A39'] = 'Test Time'
        ws['B39'] = oht_filtering_validation_result.get('test_time', 'N/A') if oht_filtering_validation_result else 'N/A'
        
        # 타겟별 결과 출력
        validation_row = 40
        if oht_filtering_validation_result:
            for target_name, target_data in oht_filtering_validation_result.items():
                if target_name in ['GL_serial', 'test_time']:
                    continue
                ws[f'A{validation_row}'] = f'{target_name} test'
                ws[f'B{validation_row}'] = 'PASS' if target_data["is_passed"] else 'FAIL'
                validation_row += 1
        else:
            ws[f'A{validation_row}'] = 'No Data Available'
            validation_row += 1

        # OHT Filtering Validation 상세 결과를 위한 새로운 시트 생성
        ws4 = wb.create_sheet('OHT test')
        
        # 상세 정보 작성 (Fail 항목만)
        detail_rows = []
        if oht_filtering_validation_result:
            for target_name, target_data in oht_filtering_validation_result.items():
                if target_name in ['GL_serial', 'test_time']:
                    continue
                for data in target_data['data']:
                    if not data['is_passed']:
                        detail_rows.append({
                            'Target': target_name,
                            'Distance': data['dist'],
                            'Device Angle': data['device_angle'],
                            'Pass/Fail': 'Fail'
                        })
        
        df_details = pd.DataFrame(detail_rows)
        
        # 상세 시트 작성
        for r in dataframe_to_rows(df_details, index=False, header=True):
            ws4.append(r)
            
        # 열 너비 자동 조정 및 스타일 적용
        for column in ws4.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            ws4.column_dimensions[column_name].width = adjusted_width
            
        # 헤더 스타일 적용
        header_font = Font(bold=True)
        for cell in ws4["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Max Distance Results Section (시작: validation_row + 2)
        ws[f'A{validation_row + 2}'] = 'Max Distance Results'
        ws[f'A{validation_row + 3}'] = 'Parameter'
        ws[f'B{validation_row + 3}'] = 'Value'
        ws[f'A{validation_row + 4}'] = 'Test Time'
        ws[f'B{validation_row + 4}'] = max_dist_result.get('test_time', 'N/A') if max_dist_result else 'N/A'
        ws[f'A{validation_row + 5}'] = 'Max Distance (9.0m) is passed?'
        ws[f'B{validation_row + 5}'] = f"{max_dist_result.get('is_passed', 'N/A')}" if max_dist_result else 'N/A'
        # 스캔각도별 감지율
        ws[f'A{validation_row + 6}'] = 'Scan Angle'
        ws[f'B{validation_row + 6}'] = 'Detection Rate'

        if max_dist_result is not None:
            # result 내부에 각 센서각도별로 정보가 저장되어 있다고 가정 (예: result["-130.0"] = {...} )
            angles = list(max_dist_result.keys())  # 예: ['-130.0', '-90.0', ...]
            max_dist_row = validation_row + 7
            
            for angle in angles:
                if angle == 'GL_serial' or angle == 'test_time' or angle == 'is_passed':
                    continue
                ws[f'A{max_dist_row}'] = f"scan angle: {angle}deg"
                ws[f'B{max_dist_row}'] = f"{max_dist_result.get(angle, {}).get('detection_ratio', 'N/A')}"
                max_dist_row += 1
        else:
            ws[f'A{validation_row + 7}'] = 'No Data Available'
            max_dist_row = validation_row + 8

        # Rear Cover Test Results Section
        rear_cover_result = inputs.get('rear_cover_result')
        ws[f'A{max_dist_row + 1}'] = 'Rear Cover Test Results'
        ws[f'A{max_dist_row + 2}'] = 'Parameter'
        ws[f'B{max_dist_row + 2}'] = 'Value'
        ws[f'A{max_dist_row + 3}'] = 'Test Time'
        ws[f'B{max_dist_row + 3}'] = rear_cover_result.get('test_time', 'N/A') if rear_cover_result else 'N/A'
        ws[f'A{max_dist_row + 4}'] = 'Test Result'
        ws[f'B{max_dist_row + 4}'] = rear_cover_result.get('result', 'N/A') if rear_cover_result else 'N/A'

        # Rear Cover Test 상세 결과를 위한 새로운 시트 생성
        ws5 = wb.create_sheet('Rear Cover Test')
        
        # 상세 정보 작성 (Fail 항목만)
        detail_rows = []
        if rear_cover_result and 'fail_datas' in rear_cover_result:
            for fail_data in rear_cover_result['fail_datas']:
                detail_rows.append({
                    'Distance (mm)': fail_data.get('test_dist', None),
                    'Target Angle (deg)': fail_data.get('test_angle', None),
                    'Device Angle (deg)': fail_data.get('test_device_angle', None),
                    'Status': 'Fail (미감지)'
                })
        
        df_details = pd.DataFrame(detail_rows)
        
        # 상세 시트 작성
        for r in dataframe_to_rows(df_details, index=False, header=True):
            ws5.append(r)
            
        # 열 너비 자동 조정 및 스타일 적용
        for column in ws5.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            ws5.column_dimensions[column_name].width = adjusted_width
            
        # 헤더 스타일 적용
        header_font = Font(bold=True)
        for cell in ws5["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2  # 여유 공간 추가
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Save the Excel file
        current_time = datetime.now().strftime("%Y%m%d_%H%M")
        report_path = os.path.join(parameters['save_path'], f"{current_time}_{GL_serial}_2nd_cal_report.xlsx")
        os.makedirs(os.path.dirname(report_path), exist_ok=True)  # 디렉토리가 없으면 생성
        wb.save(report_path)
        print(f"Report saved to: {report_path}")

        # 임시 이미지 파일 삭제
        for img_path in temp_image_files:
            try:
                if os.path.exists(img_path):
                    os.remove(img_path)
                    print(f"임시 이미지 파일 삭제됨: {img_path}")
            except Exception as e:
                print(f"이미지 파일 삭제 중 오류 발생: {str(e)}")


def run(inputs, parameters):
    try:
        tx_level_result = util_yy.load_pickle_from_zip(inputs['tx_level_result_fname'])
    except:
        tx_level_result = None
        
    try:
        home_position_result = util_yy.load_pickle_from_zip(inputs['home_position_result_fname'])
    except:
        home_position_result = None
        
    try:
        distance_test_result = util_yy.load_pickle_from_zip(inputs['distance_test_result_fname'])
    except:
        distance_test_result = None
        
    try:
        oht_filtering_table_result = util_yy.load_pickle_from_zip(inputs['oht_filtering_table_result_fname'])
    except:
        oht_filtering_table_result = None
        
    try:
        oht_filtering_validation_result = util_yy.load_pickle_from_zip(inputs['oht_filtering_validation_result_fname'])
    except:
        oht_filtering_validation_result = None
        
    try:
        max_dist_result = util_yy.load_pickle_from_zip(inputs['max_dist_result_fname'])
    except:
        max_dist_result = None

    try:
        rear_cover_result = util_yy.load_pickle_from_zip(inputs['rear_cover_result_fname'])
    except:
        rear_cover_result = None

    report_inputs = {
        'tx_level_result': tx_level_result,
        'home_position_result': home_position_result,
        'distance_test_result': distance_test_result,
        'oht_filtering_table_result': oht_filtering_table_result,
        'oht_filtering_validation_result': oht_filtering_validation_result,
        'max_dist_result': max_dist_result,
        'rear_cover_result': rear_cover_result,
    }
    
    try:
        # Call the report creation function
        create_excel_report(report_inputs, parameters)
    except ValueError as e:
        print(f"오류 발생: {str(e)}")
        return False
    return True


def unittest():
    app = QApplication(sys.argv)

    tx_level_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "TX Level 결과 파일 선택",         # 대화상자 제목
        default_parameters['tx_level_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    home_position_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "Home Position 결과 파일 선택",     # 대화상자 제목
        default_parameters['home_position_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    distance_test_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "Distance Test 결과 파일 선택",   # 대화상자 제목
        default_parameters['distance_test_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    oht_filtering_table_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "OHT Filtering Table 결과 파일 선택", # 대화상자 제목
        default_parameters['oht_filtering_table_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    oht_filtering_validation_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "OHT Filtering Validation 결과 파일 선택", # 대화상자 제목
        default_parameters['oht_filtering_validation_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    max_dist_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "Max Distance 결과 파일 선택",      # 대화상자 제목
        default_parameters['max_dist_log_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )
    rear_cover_result_fname, _ = QFileDialog.getOpenFileName(
        None,                          # 부모 위젯 없음
        "Rear Cover Test 결과 파일 선택",   # 대화상자 제목
        default_parameters['rear_cover_test_path'],             # 초기 디렉토리
        "ZIP 파일 (*.zip);;모든 파일 (*)"  # 파일 필터 (zip 우선)
    )

    inputs = {
        'tx_level_result_fname': tx_level_result_fname,
        'home_position_result_fname': home_position_result_fname,
        'distance_test_result_fname': distance_test_result_fname,
        'oht_filtering_table_result_fname': oht_filtering_table_result_fname,
        'oht_filtering_validation_result_fname': oht_filtering_validation_result_fname,
        'max_dist_result_fname': max_dist_result_fname,
        'rear_cover_result_fname': rear_cover_result_fname,
    }
    parameters = default_parameters

    run(inputs, parameters)
    app.quit()


if __name__ == '__main__':
    unittest()
    print('Done')